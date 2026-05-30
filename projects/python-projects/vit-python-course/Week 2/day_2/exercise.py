# ITP Week 2 Day 2 Exercise

# import the appropriate method from the correct module
import openpyxl

# Import the workbook that you updated in today's practice from
wb = openpyxl.load_workbook(r'C:\Users\amesb\Desktop\WebDevTraining\TheCarpetBag\inventory.xlsx')

# access and store the appropriate worksheet to the variable 'ws'
ws=wb['CURRENT_MONTH_INVENTORY']

# Define a function called add_order_amount that takes in a single parameter called 'row'
def add_order_amount(row):#row will be the row numb in excel
    #pull values FROM the sheet using the row number
    max_amount = ws["C" + str(row)].value
    reorder_threshold = ws["D" + str(row)].value
    quantity = ws["E" + str(row)].value

    # IF the quantity is less or equal to than the threshold,
    if quantity <= reorder_threshold:
        # than calculate the order_amount (max_amount - quantity) 
        order_amount=max_amount - quantity
        # assign the value to that row, column 6
        ws["F" + str(row)]=order_amount #stores the result into the new column and rows
         
# TIP: create variables for quantity, threshold, max_amount that retrieves the values first for cleanliness

# Perform a for..in loop through the range(2, len(inventory.rows))
for row in range(2, ws.max_row + 1): #start with row 2-14+1 to get 15

#   - call the function add_order_amount() passing in the number of the range
    add_order_amount(row)
    
    
wb.save("inventory.xlsx")