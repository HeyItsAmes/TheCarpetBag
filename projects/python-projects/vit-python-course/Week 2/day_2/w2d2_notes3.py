#Import the appropriate Library

from openpyxl import load_workbook #tool to access a previously stored file, opens exsisting file on computer

#load previous workbook

wb=load_workbook("student_report.xlsx") #wb is a python object that represents of the excel file in memory so u can manipulate it

#Look at sheetnames - hardcode version
print(wb.sheetnames) #You need to always know what you're working with 

#Look at sheetnames -  Loop through the sheets version
for sheet in wb:
    print(sheet.title)
    
#Select a worksheet
ws=wb["Student Grades"]  #Go into the workbook and get me the worksheet that's called Student Grades

#Look at all rows as cell objects

all_rows=ws.rows #With ws.rows we don't get the values, we only get cell objects here. It just would return A1, not the values in the cell
print(list(all_rows)) #Now you get the data in the cell, text, forumlas, any formatting, etc

all_values = ws.values #prints out the values inside the cells
print(list(all_values))

#You can also loop through row by row - each row of values - and get a clean organized output
for row in ws.values:
    print(row)
    
#Loop through each row of values example 2
count=0

for row in ws.values:
    print(row)
    count +=1
print("Total rows: ", count) #gives you total rows

# if you had a big dataset, limit the amount to like 5 rows
# print("Total rows: ", count)[5] #limits the amount of rows returned in output

#if you want to limit data to the most recent "youthful data"
loaded_ws= wb["Student Grades"]
for row in loaded_ws.values:
    #Skip rows where every cell is empty
    if any(cell is not None for cell in row):
        print(row) #This took away all the none, none, none row with no info in cells
        
#Slicing to limit which data you want
loaded_ws=wb["Student Grades"]
rows=list(loaded_ws.values)
student_rows=rows[0:4]
for row in student_rows:
    print(row)



