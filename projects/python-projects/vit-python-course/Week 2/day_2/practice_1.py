# ITP Week 2 Day 2 (In-Class) Practice 1
# 
# You will continue to work on the inventory spreadsheet that you created from yesterday's exercise
# import the appropriate method from the correct module

import openpyxl
from openpyxl import Workbook

import os
# print(os.getcwd())  #print the current working directory

# Import the workbook that you created in yesterday's exercise from
wb = openpyxl.load_workbook(r'C:\Users\amesb\Desktop\WebDevTraining\TheCarpetBag\inventory.xlsx')
# print(type(wb)) 
# print(wb.sheetnames)

# verify what sheet names are available
names = wb.sheetnames
# print(names)

# access and store the appropriate worksheet to the variable 'ws'
ws=wb['CURRENT_MONTH_INVENTORY']

# Print out the cell values for each row
all_rows=ws.rows #With ws.rows we don't get the values, we only get cell objects here. It just would return A1, not the values in the cell
print(list(all_rows)) #Now you get the data in the cell, text, forumlas, any formatting, etc


# Create a new column within that worksheet called order_amount
#Header for row 1 column F
ws["F1"] = "order_amount"

# save the latest changes
wb.save("inventory.xlsx")