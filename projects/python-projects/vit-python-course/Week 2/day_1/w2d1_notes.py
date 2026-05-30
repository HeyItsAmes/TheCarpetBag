#Import Workbook from the openpyxl library
#Workbook is used to create a new excel file
import openpyxl
from openpyxl import Workbook #Workbook must be capitolized

#Create a brand new excel file
#Think of it as opening a blank excel file
wb= Workbook()  

#Get the default Worksheet that comes with every new workbook
#A worksheet is one tab inside of the excel file
ws = wb.active #This is the Sheet1 Worksheet tab in a Workbook

#Rename the default worksheet tab from "Sheet" to "Student Grade"
ws.title="Student Grades"

#Create a new worksheet tab called "Attendance"
wb.create_sheet("Attendance")

#Create another worksheet tab called "Homework"
wb.create_sheet("Homework")

#Print all the names of the worksheet in the workbook.
print(wb.sheetnames)

#METHOD 1 Writing values using A1(excel) Style - how to write to cells first way
#Great for headers, or one cell - easiest to read bcuz it matches Excel.  good for when u know the cell.
ws["A1"]="Name"
ws["B1"]="Grade"

#METHOD 2 Worksheet Dot Cell Method - second way to write to cells with row/column
#or instead of excel style use Row Col numbers
#Great for LOOPS or lots of data
#Write "Amy" into row 2, column 1 is Column A so this is A2
ws.cell(row=2, column=1, value="Amy")

#write 100 to row 2, column 2, Column 2 is Column B so this cell B2
ws.cell(row=2, column=2, value=100)

#write Sarah into row 3, column 1
#This is cell A3
ws.cell(row=3, column=1, value="Sarah")

#write 95 into row 3, column 2
#This is cell B3
ws.cell(row=3, column=2, value=95)

#METHOD 3 - Using iter_rows() - Go thru ws row by row, it gives u a row of cells
#Each cell in excel is an object in openpyxl
#OpenpyXl you'll only get cells once you create them in memory

#Fill a 3x3 block w numbers
#Create a number variable starting it at 1 - excel starts counting at 1 but python starts at zero so create a variale so python won't get confused bcuz we have to start counting at one for excel
number=1 

#iter_rows to loop through a rectangular block of cells
for row in ws.iter_rows(min_row=5, max_row=7, min_col=1, max_col=3):
        #EAch row contains multiple cells
        #This innter loop visits each cell inside the current row 
        for cell in row: 
            #put the current number into the cell
            cell.value=number
            #increase the number by 1 so the next cell gets the next number
            number += 1
      
#Two ways to save:      
# wb.save("spreadsheets/w2d1_lecture.xlsx")
wb.save("w2d1_lecture.xlsx")



