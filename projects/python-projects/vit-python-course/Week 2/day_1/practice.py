# ITP Week 2 Day 1 (In-Class) Practice

# A1. from the appropriate library, import only the Workbook
import openpyxl
from openpyxl import Workbook

# A2. Before anything, we need a workbook to work with..
wb= Workbook()

# A3. We need to interact with a single worksheet.
ws = wb.active
# print(wb.sheetnames)

# A4. assign the value of "First Name" to A1
ws["A1"]="First Name"

# A5. assign the value of "Last Name" to B1
ws["B1"]="Last Name"

# STOP HERE - RETURN TO LECTURE

# B1. For all of column A, starting at row 2 until row 10, make the cell values: "Gabriel" (attempt a loop)
        
for row_num in range(2, 11):  #will produce 2, 3, 4, 5, 6, 7,...10
    ws["A" + str(row_num)]="Gabriel" # Using ex w/ first name above, this will use column A only, specifiy rows 2-10 only, combines A + row number and assigns Gabriel to each one.


last_names = ['Rolley', 'Smith', 'Balenga', 'Issac', 'Cruise', 'Depp', 'Heard', 'Qiao', 'Biden']

# B2. Loop through a range from row 2 to 10 and assign the cell value to last names according to index in column B
# NOTE: PAY ATTENTION to the starting number of the range and how it differs from the starting index of the list.
# #start with index zero for Rolley for python but Rolly will be going into row 2 in excel

list_position=0 #this will get updated each loop because of +=1
for row_num in range(2, 11):  #will produce 2, 3, 4, 5, 6, 7,...10
    ws["B" + str(row_num)]=last_names[list_position] #Targets col B, rows 2-10, assigns list pos
    list_position += 1 #moves index fwd by 1 each loop by inscreasing index number
    


# B3. Save the file
wb.save("day_1_practice.xlsx")