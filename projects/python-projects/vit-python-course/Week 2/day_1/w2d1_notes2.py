# Week 2 Day 1 
# Will be doing day 2 all practice files in class

# import proper library
from openpyxl import Workbook

#Create a new workbook()
wb = Workbook()

#Access the default worksheet
ws = wb.active
 
#Rename the worksheet
ws.title="Student Grades"

#Create additional worksheets
attendance_sheet=wb.create_sheet("Attendance")
homework_sheet=wb.create_sheet("Homework")

#Print all worksheet names
# print(wb.sheetnames)

#Student Grade Section
#Create Headers, Excel Style "A1"

ws["A1"] = "Name"
ws["B1"] = "Quiz 1"
ws["C1"] = "Quiz 2"
ws["D1"] = "Final Grade"

#Add Student data
#Creates a list so it represents a row similar to Excel
students =[
        ["Jael", 95, 100],
        ["Mike", 88, 92], 
        ["David", 91, 85],
    ] 

#Start writing student at row 2
current_row=2

for student in students: #for ea student inside student list this is what we want to happen..
    name = student[0] #Jael
    quiz1 = student[1]  #95
    quiz2 = student[2] #100
    
    #Calculate the final grade
    final_grade=(quiz1 + quiz2) /2 #getting an average of the two quizes. for that student, calculate final grade
    
    #Then go thru and Write data into excell cell
    ws.cell(row=current_row, column=1, value=name) #first loop, writes name for Jael
    ws.cell(row=current_row, column=2, value=quiz1)#first loop, writes quiz for Jael
    ws.cell(row=current_row, column=3, value=quiz2)#first loop, writes quiz for Jael
    ws.cell(row=current_row, column=4, value=final_grade)#first loop, writes final grade for Jael
    
    current_row +=1 # then goes to next student
    
    #Multiplucation table
    
    #Add Title Section
    ws["A6"] = "Multiplication Table" 
    
    #3x3 Multiplication Table (nested loop)
    
    for row in range(1, 4): #stops before 4, 1-3 basically
        for col in range(1, 4): #stops before 4
            value=row * col #multipily row times colum 1x1, 1x2, 1x3, etc

            ws.cell(row=row + 6, column=col, value=value) #row + 6 means start my tble at lower part of sheet so table will start at row 7
            
    #Save workbook
    wb.save("student_report.xlsx")
    print("You did it!")