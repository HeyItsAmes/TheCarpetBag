# ITP Week 2 Day 1 Exercise

# SCENARIO: You're a store owner receiving the inventory report for this month.
# You will receive the product order spreadsheet soon and it is easier to calculate your order
# if your inventory was also on a spreadsheet! Recreate the following spreadsheet with Python: 

# don't forget your appropriate imports.
import openpyxl
from openpyxl import Workbook

wb= Workbook()

# assign the title of the initial active sheet to "CURRENT_MONTH_INVENTORY"
ws = wb.active
ws.title="CURRENT_MONTH_INVENTORY"

#   product_name    product_id  max_amount      reorder_threshold   quantity
# 	oreo            2323        1000            300                 743
# 	coke            6545        500             100                 101
# 	pepsi	        3456        200             50                  137
# 	lays_chip	    4567        1500            500                 364
# 	pringles	    2134        2000            600                 120
# 	sour_worms	    2362        100             10                  85
# 	choco_cookies   0923	    200             25                  24
# 	donuts	        2786        200             25                  12
# 	hot_dogs	    6723        100             10                  39
# 	ice_cream	    9237        200             50                  234
# 	gum	            2092        3500            1000                1232
# 	pretzels        8246	    100             5                   11
# 	kit_kat	        9276        1000            250                 249

# TIP: create a list of each column (ie. product_names = [oreo, ...]) and use those to loop through :)
#List Creation for each column
product_names = [   "oreo",
    "coke",
    "pepsi",
    "lays_chip",
    "pringles",
    "sour_worms",
    "choco_cookies",
    "donuts",
    "hot_dogs",
    "ice_cream",
    "gum",
    "pretzels",
    "kit_kat"]

product_id = [
    2323,
    6545,
    3456,
    4567,
    2134,
    2362,
    923,
    2786,
    6723,
    9237,
    2092,
    8246,
    9276
]
max_amount = [
    1000,
    500,
    200,
    1500,
    2000,
    100,
    200,
    200,
    100,
    200,
    3500,
    100,
    1000
]

reorder_threshold = [
    300,
    100,
    50,
    500,
    600,
    10,
    25,
    25,
    10,
    50,
    1000,
    5,
    250
]

quantity = [
    743,
    101,
    137,
    364,
    120,
    85,
    24,
    12,
    39,
    234,
    1232,
    11,
    249
]

#Headers for row 1
ws["A1"] = "product_name"
ws["B1"] = "product_id"
ws["C1"] = "max_amount"
ws["D1"] = "reorder_threshold"
ws["E1"] = "quantity"

#Build the rest of the excel rows and columns

#Column A
product_index = 0
for row_num in range(2, 15): #will produce 2, 3, 4, 5, 6, 7,...10
    ws["A" + str(row_num)] = product_names[product_index]  # Targets column A, specifiy rows 2-15, combines A + row number and assigns product name and its position each pass.
    product_index += 1  
    
    
#Column B
product_index = 0
for row_num in range(2, 15):
    ws["B" + str(row_num)] = product_id[product_index]
    product_index += 1
    
#Column C
product_index = 0
for row_num in range(2, 15):
    ws["C" + str(row_num)] = max_amount[product_index]
    product_index += 1
    
#Column D
product_index = 0
for row_num in range(2, 15):
    ws["D" + str(row_num)] = reorder_threshold[product_index]
    product_index += 1
    
#Column E
product_index = 0
for row_num in range(2, 15):
    ws["E" + str(row_num)] = quantity[product_index]
    product_index += 1
    
    

# save your file
wb.save("inventory.xlsx")