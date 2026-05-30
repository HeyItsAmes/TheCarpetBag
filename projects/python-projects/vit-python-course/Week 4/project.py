# cache_me_outside team: Blake La Touf, Craig Guffey, Irina La Touf, Carl Daniels, Amy Oglesbee

# Import the required modules
# Import requests to get data from websites/API
# Import JSON to convert JSON text into Python data
import requests
import json

# Use workbook to create a new excel spreadsheet
from openpyxl import Workbook

# API CALL #1 (states)
# Get data from this url
state_url = "https://state-data-api.proxy.beeceptor.com/state-details"
# Send a GET request to the API - get data from the API/website"
state_response = requests.get(state_url)
# print(state_response)  #200 verifies success

# API CALL #2 (travel)
# get data from this URL
travel_url = "https://state-data-api.proxy.beeceptor.com/state-travel-data"
# Send a GET request to the API - get data from the API/website"
travel_response = requests.get(travel_url)
# print(travel_response) #200 verifies success

# convert JSON string into a Python object (a list of dictionaries) and store
state_data = json.loads(state_response.text)
# Print first to understand the structure
# print(type(state_data))  #type of data is a list

# print(
#     state_data[0]
# )  # Output: {'state': 'Alabama', 'region': 'South', 'population': 5108468}   NOT A LIST, gives 1st index/item vibes


travel_data = json.loads(travel_response.text)
# print(type(travel_data)) #type of data is a list
# print(travel_data[:1])   # Output: [{'state': 'Alabama', 'top_attraction': 'Gulf Shores', 'average_trip_cost': 780}]

# Pretty-print to the console cuz no one can read that ish
# print(json.dumps(state_data, indent=4))
# print(json.dumps(travel_data, indent=4))
# conclusion: both datasets have a common key called "state"

# merge datasets by "state" step by step:
combined_data = []  # variable waiting to receive it's goodies (list of dictionaries)
# Take one state record, then search every travel record looking for a matching state name.
for state in state_data:
    for travel in travel_data:
        # print(state["state"], travel["state"])  #output: comparison of both state columns

        # IF statement so it only does something with state names that match
        if state["state"] == travel["state"]:
            # if state and travel names are samezies then print it
            # print(state["state"], "DING! DING! ITS A MATCH")

            # merge matched data and store it in the variable combined_record
            combined_record = {
                "state": state["state"],
                "region": state["region"],
                "population": state["population"],
                "top_attraction": travel["top_attraction"],
                "average_trip_cost": travel["average_trip_cost"],
            }
            combined_data.append(combined_record)

# print(combined_data[1])  #nailed it

# Check if 50 states merged:
# print(len(combined_data)) #50 also nailed it


# build Excel workbook
wb = Workbook()

# write sheet 1
ws = wb.active  # Get the default worksheet
ws.title = "State Travel Report"  # Rename the default sheet

# Headers
ws.append(["State", "Region", "Population", "Top Attraction", "Average Trip Cost"])

for data in combined_data:
    ws.append(
        [
            data["state"],
            data["region"],
            data["population"],
            data["top_attraction"],
            data["average_trip_cost"],
        ]
    )


# write sheet 2 (grouped by region)
ws2 = wb.create_sheet("States by Region")
# headers_region = ["State", "Top Attraction", "Average Trip Cost", "Region"]

# for col_num, header in enumerate(headers_region, 1):
#     ws2.cell(row=1, column=col_num, value=header)


# ordering sheet 2 by the region column

# for record in combined_data:
#     print(record["state"], record["region"])  #print to check the state and region columns

# Ordered list will control the report structure during output

ws2.append(["State", "Top Attraction", "Average Trip Cost", "Region"])

regions = ["Midwest", "West", "South", "Northeast"]

for region in regions:
    for record in combined_data:
        if record["region"] == region:
            ws2.append(
                [
                    record["state"],
                    record["top_attraction"],
                    record["average_trip_cost"],
                    record["region"],
                ]
            )
    ws2.append([])

# Fancy auto-formatting
# Freeze top row on both sheets
ws.freeze_panes = "A2"
ws2.freeze_panes = "A2"

# Auto-fit column widths for sheet 1
for column in ws.columns:
    max_length = max(len(str(cell.value)) for cell in column if cell.value)
    ws.column_dimensions[column[0].column_letter].width = max_length + 2

# Auto-fit column widths for sheet 2
for column in ws2.columns:
    max_length = max(len(str(cell.value)) for cell in column if cell.value)
    ws2.column_dimensions[column[0].column_letter].width = max_length + 2


# save that file with some fancy footwork
wb.save("state_travel_report_formatted.xlsx")
print(
    "\033[1;95mIf you're seeing this message Python and cache_me_outside finally made it to the end without console meltdowns.\033[0m"
)
print("\033[1;95m₍^. .^₎⟆\033[0m")
