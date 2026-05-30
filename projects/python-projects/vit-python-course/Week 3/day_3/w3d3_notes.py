# Import approriate modules
# Import requests to get data from the internet/API
# Import JSON to convert JSON text into Python data
import requests
import json

# Use workbook to create a new excel spreadsheet
from openpyxl import Workbook

# Get data from API
url = "https://jsonplaceholder.typicode.com/posts"
# Send a GET request to the API
# "Give Data from the API or from url"
response = requests.get(url)
# Verify sucess
print(response)
# response.text - raw JSON data as a string
data = json.loads(response.text)
print(type(data))
# Print first post to understand the structure
print(data[:1])

# Create Spreadsheet

# create a new Excel Workbook
wb = Workbook()

# Select active worksheet
ws = wb.active

# Rename the worksheet
ws.title = "Posts"

# Add column headers
for col, key in enumerate(data[0].keys(), 1):
    # write each key into row 1
    ws.cell(row=1, column=col, value=key)

# Write API data
for row, post in enumerate(data[5], 2):
    for col, value in enumerate(post.values(), 1):
        ws.cell(row=row, column=col, value=str(value))

# Save workbook
wb.save("post.xlsx")
print("Created my first excell workbook using API")
