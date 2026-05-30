# ITP Week 3 Day 3 Exercise

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation
# we want to make a copy of the Rick and Morty database (which is provided through the api)
# EASY MODE

# import the appropriate modules (you have 3)

from openpyxl import Workbook  #create excel file
import requests #used to contact websites/APIs
import json #parses JSON


character_url = "https://rickandmortyapi.com/api/character"  #where the data comes from
# set up a workbook and worksheet titled "Rick and Morty Characters"

wb=Workbook() #new excel file
ws=wb.active  #get the first sheet, the default
ws.title="Rick and Morty Characters"  #renames the tab in excel

# # assign a variable 'data' with the returned GET request

response=requests.get(character_url) #telling API to send me the character data (placing an order)
json_data=response.text  #API will send back JSON (delivery of order)
data=json.loads(json_data) #convert JSON to Python (shipping label and tracking info)
# print(data)  The results are the contents inside the box
# print(json.dumps(data, indent=2))

# create the appropriate headers in openpyxl for all of the keys for a single character
#SYNTAX  data["results"[0]]
character = (data["results"][0]) #grabbing the first character, from the results list
# print(character.keys()) #print the fields from the dictionary
#now the keys can be used for headers in an excel sheet

# loop through all of the 'results' of the data to populate the rows and columns for each character

# NOTE: due to the headers, the rows need to be offset by one!

headers=character.keys() #grabs headers

for col, header in enumerate(headers, start=1): #gives headers column numbers
    ws.cell(row=1, column=col).value=header #puts headers into row 1

# MEDIUM MODE

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"

ws_locations = wb.create_sheet("Rick and Morty Locations")

ws_episodes = wb.create_sheet("Rick and Morty Episodes")

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)
# 
# 
response_locations = requests.get("https://rickandmortyapi.com/api/location")
json_locations = response_locations.text
locations_data = json.loads(json_locations)

response_episodes = requests.get("https://rickandmortyapi.com/api/episode")
json_episodes = response_episodes.text
episodes_data = json.loads(json_episodes)
episodes=episodes_data["results"]

# populate the new worksheets appropriately with all of the data!

locations = locations_data["results"]

print(locations[0])

# NOTE: don't forget your headers!

headers = locations[0].keys()

for col, header in enumerate(headers, start=1):
    ws_locations.cell(row=1, column=col).value = header

# HARD MODE
# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)

for row_index, location in enumerate(locations, start=2):
    for col_index, key in enumerate(location.keys(), start=1):
        value = location[key]

        if isinstance(value, list):
            value = ", ".join(value)

        ws_locations.cell(row=row_index, column=col_index).value = value
        
headers = episodes[0].keys()

for col, header in enumerate(headers, start=1):
    ws_episodes.cell(row=1, column=col).value = header
    
for row_index, episode in enumerate(episodes, start=2):
    for col_index, key in enumerate(episode.keys(), start=1):
        value = episode[key]

    if isinstance(value, list):
        value = ", ".join(value)

ws_episodes.cell(row=row_index, column=col_index).value = value

# NIGHTMARE
# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url
# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)


wb.save("exercise.xlsx")
