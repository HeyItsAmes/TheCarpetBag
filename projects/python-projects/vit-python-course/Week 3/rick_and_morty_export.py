# Rick and Morty API Data Export Project 
# Week 3 ViT Python Fundamentals Bootcamp

from openpyxl import Workbook
import requests
import json

# API endpoint for character data
character_url = "https://rickandmortyapi.com/api/character"

# Create workbook and first worksheet
wb = Workbook()
ws = wb.active
ws.title = "Rick and Morty Characters"

# Retrieve character data from the API
response = requests.get(character_url)
json_data = response.text
data = json.loads(json_data)

# Use the first character record to generate worksheet headers
character = data["results"][0]

headers = character.keys()

for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col).value = header

# Create worksheets for locations and episodes
ws_locations = wb.create_sheet("Rick and Morty Locations")
ws_episodes = wb.create_sheet("Rick and Morty Episodes")

# Retrieve location data
response_locations = requests.get("https://rickandmortyapi.com/api/location")
json_locations = response_locations.text
locations_data = json.loads(json_locations)

# Retrieve episode data
response_episodes = requests.get("https://rickandmortyapi.com/api/episode")
json_episodes = response_episodes.text
episodes_data = json.loads(json_episodes)

locations = locations_data["results"]
episodes = episodes_data["results"]

# Create headers for Locations worksheet
headers = locations[0].keys()

for col, header in enumerate(headers, start=1):
    ws_locations.cell(row=1, column=col).value = header

# Populate Locations worksheet
for row_index, location in enumerate(locations, start=2):
    for col_index, key in enumerate(location.keys(), start=1):
        value = location[key]

        if isinstance(value, list):
            value = ", ".join(value)

        ws_locations.cell(row=row_index, column=col_index).value = value

# Create headers for Episodes worksheet
headers = episodes[0].keys()

for col, header in enumerate(headers, start=1):
    ws_episodes.cell(row=1, column=col).value = header

# Populate Episodes worksheet
for row_index, episode in enumerate(episodes, start=2):
    for col_index, key in enumerate(episode.keys(), start=1):
        value = episode[key]

        if isinstance(value, list):
            value = ", ".join(value)

        ws_episodes.cell(row=row_index, column=col_index).value = value

# Future Enhancement:
# The API only returns the first page of results.
# Characters worksheet only writes headers (as directed by Instructor).  I Could finish the #character row export section.

wb.save("exercise.xlsx")

#Running this file will output:
#Character sheet (headers only)
#Location sheet (populated)
#Episode sheet (populated)